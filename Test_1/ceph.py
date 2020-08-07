import os
import time
import openpyxl
# import datetime
# import boto3
# import boto.s3.connection
from boto3.session import Session

access_key = 'S7BQDCHX99G7H3YDX0EX'
secret_key = 'w6rnLjMp529FRrtb3Ds6YE51sTjiwe6sALAdOYuS'

excel_client = openpyxl.Workbook()
excel_sheet1 = excel_client.create_sheet(title="文件上传性能测试")
excel_sheet2 = excel_client.create_sheet(index=1, title="文件下载性能测试")
excel_row = ['A', 'B', 'C', 'D']
excel_sheet1['A1'] = "对象服务"
excel_sheet1['B1'] = "0.7 MB"
excel_sheet1['C1'] = "5.88 MB"
excel_sheet1['D1'] = "88.9 MB"

excel_sheet2['A1'] = "对象服务"
excel_sheet2['B1'] = "0.7 MB"
excel_sheet2['C1'] = "5.88 MB"
excel_sheet2['D1'] = "88.9 MB"

# startTime2 = time.time()
# for bucket in conn.get_all_buckets():
#     print("{name}\t{created}".format(name=bucket.name, created=bucket.creation_date))
# endTime2 = time.time()
# print("date: {0} ms".format((endTime2 - startTime2) * 1000))


# startTime = datetime.datetime.now()
# for bucket in conn.get_all_buckets():
#     print("{name}\t{created}".format(name=bucket.name, created=bucket.creation_date))
# endTime = datetime.datetime.now()
# print(endTime - startTime)


session = Session(aws_access_key_id=access_key, aws_secret_access_key=secret_key)
url = "http://192.168.8.106:7480"
ceph_client = session.client('s3', endpoint_url=url)
ceph_resource = session.resource('s3', endpoint_url=url)

objects_list = ["一份脱敏记录.xml", "最大入院记录.xml", "88.9MB.zip"]

for number in range(2, 102):
    list_number = 1
    for object_name in objects_list:
        startTime = time.time()
        ceph_resource.meta.client.upload_file("d:\\{0}".format(object_name), "test1", object_name)
        endTime = time.time()
        TimeConsuming = round((endTime - startTime) * 1000, 0)
        # print("Upload {1} File Time: {0} ms".format((endTime - startTime) * 1000, object_name))
        excel_sheet1['A{0}'.format(number)] = "Ceph_Python_{0}".format(number - 1)
        excel_sheet1['{0}{1}'.format(excel_row[list_number], number)] = TimeConsuming
        print('{0}{1}'.format(excel_row[list_number], number))
        list_number += 1


print()
# response = ceph_client.list_objects(Bucket='test1')
# for tmp in response['Contents']:
#     print("Object: {0}".format(tmp['Key']))

for number in range(2, 102):
    list_number = 1
    for object_name in objects_list:
        startTime = time.time()
        download_fle = ceph_resource.meta.client.download_file("test1", object_name, "d:\\tmp\\{0}".format(object_name))
        endTime = time.time()
        TimeConsuming = round((endTime - startTime) * 1000, 0)
        # print("Download {1} File Time: {0} ms".format((endTime - startTime) * 1000, object_name))
        excel_sheet2['A{0}'.format(number)] = "Ceph_Python_{0}".format(number - 1)
        excel_sheet2['{0}{1}'.format(excel_row[list_number], number)] = TimeConsuming
        print('{0}{1}'.format(excel_row[list_number], number))
        list_number += 1

excel_client.save('d:\\tmp\\ceph_tmp.xlsx')
excel_client.close()

for object_name in objects_list:
    ceph_client.delete_object(Bucket="test1", Key=object_name)

try:
    response = ceph_client.list_objects(Bucket='test1')
    for tmp in response['Contents']:
        print("Delete to Object: {0}".format(tmp['Key']))
except Exception:
    print("List Object Null")

for object_name_tmp in objects_list:
    os.remove("D:\\tmp\\{0}".format(object_name_tmp))
