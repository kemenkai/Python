import os
import time
import openpyxl
from minio import Minio
from minio.error import ResponseError

minio_client = Minio("192.168.8.106:9000",
                     access_key="AKIAIOSFODNN7EXAMPLE",
                     secret_key="wJalrXUtnFEMI/K7MDENG/bPxRfiCYEXAMPLEKEY",
                     secure=False)

excel_client = openpyxl.Workbook()
excel_sheet1 = excel_client.create_sheet(index=0, title="上传文件测试结果")
excel_sheet2 = excel_client.create_sheet(index=1, title="下载文件测试结果")
excel_row = ['A', 'B', 'C', 'D']
excel_sheet1['A1'] = "对象服务"
excel_sheet1['B1'] = "0.7 MB"
excel_sheet1['C1'] = "5.88 MB"
excel_sheet1['D1'] = "88.9 MB"

excel_sheet2['A1'] = "对象服务"
excel_sheet2['B1'] = "0.7 MB"
excel_sheet2['C1'] = "5.88 MB"
excel_sheet2['D1'] = "88.9 MB"


def upload(object_name, file_path):
    try:
        minio_client.fput_object("test1", object_name, file_path, content_type="application/text")
    except ResponseError as err:
        print(err.message)


def remove_object(objects):
    try:
        for del_err in minio_client.remove_objects("test1", objects):
            print("Remove Error: {}".format(del_err))
    except ResponseError as err:
        print(err.message)


def download_object(object_name, file_path):
    # Get a full object and prints the original object stat information.
    try:
        minio_client.fget_object('test1', object_name, file_path)
    except ResponseError as err:
        print(err.message)


objects_list = ["一份脱敏记录.xml", "最大入院记录.xml", "88.9MB.zip"]
for number in range(2, 102):
    list_number = 1
    for object_name_tmp in objects_list:
        startTime = time.time()
        upload(object_name_tmp, "D:\\{0}".format(object_name_tmp))
        endTime = time.time()
        TimeConsuming = (endTime - startTime) * 1000
        # print("Upload {1} File Time: {0} ms".format(TimeConsuming, object_name_tmp))
        excel_sheet1['A{0}'.format(number)] = "Minio_Python_{0}".format(number - 1)
        excel_sheet1['{0}{1}'.format(excel_row[list_number], number)] = round(TimeConsuming, 0)
        print('{0}{1}'.format(excel_row[list_number], number))
        list_number += 1

print("\n")
list_objects = minio_client.list_objects("test1")
# for tmp in list_objects:
#     print("Bucket Name: {0}, Object Name: {1}".format(tmp.bucket_name, tmp.object_name))
# print("\n")
for number in range(2, 102):
    list_number = 1
    for object_name_tmp in objects_list:
        startTime = time.time()
        download_object(object_name_tmp, "d:\\tmp\\{0}".format(object_name_tmp))
        endTime = time.time()
        TimeConsuming = (endTime - startTime) * 1000
        # print("Download {0} File Time: {1} ms".format(object_name_tmp, (endTime - startTime) * 1000))
        excel_sheet2["A{0}".format(number)] = "Minio_Python_{0}".format(number - 1)
        excel_sheet2["{0}{1}".format(excel_row[list_number], number)] = round(TimeConsuming, 0)
        print('{0}{1}'.format(excel_row[list_number], number))
        list_number += 1

for object_name_tmp in objects_list:
    os.remove("D:\\tmp\\{0}".format(object_name_tmp))
remove_object(objects_list)

excel_client.save("d:\\tmp\\minio_tmp.xlsx")
excel_client.close()
